"""Lectura de listas de referencia (analistas, areas, causas) y de la linea
base de SLA (numero de dispositivos / horas del mes) desde la Planilla S@MI,
y escritura de un evento nuevo capturado en el formulario.

Nota importante: las celdas "Tiempo Total / Tiempo Minutos / Tiempo en
Horas" de cada hoja son formulas que Excel recalcula al abrir el archivo.
openpyxl no ejecuta formulas, asi que este modulo nunca lee esas celdas
para calcular disponibilidad: recalcula la duracion de cada evento en
Python directamente a partir de Fecha/Hora Inicio y Fecha/Hora Fin, que
son valores literales y siempre estan disponibles."""

from datetime import datetime

import openpyxl
from openpyxl.utils import column_index_from_string

from schema import FIRST_DATA_ROW, HEADER_ROW, PLATFORMS, RESUMEN_ROWS, PlatformSchema
from validation import EventoPayload

MAX_ROW_SCAN = 5000


def load_workbook_for_write(path: str):
    """Carga el libro para edicion. Las imagenes incrustadas (logo/decoracion
    en 'Resumen') se descartan: openpyxl consume el flujo de la imagen la
    primera vez que se llama a wb.save(), y una segunda llamada sobre el
    mismo objeto (por ejemplo, cada vez que Streamlit vuelve a renderizar la
    pagina para armar la descarga) revienta con
    'ValueError: I/O operation on closed file'. Sin las imagenes, wb.save()
    se puede invocar cuantas veces haga falta sin ese riesgo."""
    wb = openpyxl.load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        ws._images = []
    return wb


def clean_analista_name(raw: str) -> str:
    """La planilla guarda el analista como 'Nombre / turno' (p. ej.
    'Anderson Garcia Cataño / 06:00 - 14:00'). El formulario y la revision
    solo muestran y guardan el nombre: el turno no aporta nada al calculo de
    SLA y solo ensucia la lista de analistas (la misma persona aparece una
    vez por cada turno que cubre)."""
    if not raw:
        return ""
    return str(raw).split("/")[0].strip()


def load_lookups(wb) -> dict:
    """Extrae Analistas, Areas y Causas desde la hoja 'Hoja1'."""
    ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb["Hoja1 (2)"]

    analistas: list[str] = []
    areas: list[str] = []
    causas: list[str] = []

    causa_header_seen = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        col_c = row[2] if len(row) > 2 else None
        col_g = row[6] if len(row) > 6 else None

        if col_g and col_g != "AREAS":
            areas.append(str(col_g).strip())

        if col_c == "CAUSA":
            causa_header_seen += 1
            continue
        if col_c:
            text = str(col_c).strip()
            if causa_header_seen > 0:
                if text not in causas:
                    causas.append(text)
            elif "/" in text:  # filas de analista tienen el formato "Nombre / turno"
                nombre = clean_analista_name(text)
                if nombre and nombre not in analistas:
                    analistas.append(nombre)

    return {
        "analistas": sorted(analistas, key=str.lower),
        "areas": sorted(set(areas), key=str.lower),
        "causas": sorted(set(causas), key=str.lower),
    }


def load_resumen_baseline(wb, platform: str) -> dict | None:
    """Retorna {'dispositivos', 'dias_mes', 'total_horas_mes'} para la
    plataforma, o None si la hoja Resumen no tiene fila para ella (caso de
    'Data Center', que no aparece en el resumen del archivo de origen)."""
    if platform not in RESUMEN_ROWS or "Resumen" not in wb.sheetnames:
        return None
    ws = wb["Resumen"]
    row = RESUMEN_ROWS[platform]
    dispositivos = ws.cell(row=row, column=5).value  # columna E
    dias_mes = ws["D3"].value
    if not isinstance(dispositivos, (int, float)) or not isinstance(dias_mes, (int, float)):
        return None
    return {
        "dispositivos": dispositivos,
        "dias_mes": dias_mes,
        "total_horas_mes": dispositivos * dias_mes * 24,
    }


def _col(letter: str) -> int:
    return column_index_from_string(letter)


def find_next_row(ws, schema: PlatformSchema) -> tuple[int, int | None]:
    """Busca la primera fila 'placeholder' (Item con numero, dispositivo
    vacio) para reutilizar. Si no encuentra ninguna, retorna la fila
    siguiente a la ultima usada para crear una fila nueva.
    Retorna (fila, ultimo_item_numerico_visto)."""
    cols = schema.columns
    item_col = _col(cols["item"])
    device_col = _col(cols["device"])

    last_item_row = HEADER_ROW
    last_item_value = 0
    for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + MAX_ROW_SCAN):
        item_value = ws.cell(row=row, column=item_col).value
        device_value = ws.cell(row=row, column=device_col).value
        if item_value in (None, ""):
            break
        last_item_row = row
        if isinstance(item_value, (int, float)):
            last_item_value = int(item_value)
        if device_value in (None, ""):
            return row, last_item_value

    return last_item_row + 1, last_item_value


def append_event(wb, platform: str, payload: EventoPayload) -> int:
    """Escribe el evento validado en la primera fila disponible de la hoja
    de la plataforma. Retorna el numero de fila usado."""
    schema = PLATFORMS[platform]
    ws = wb[schema.sheet]
    cols = schema.columns
    row, last_item_value = find_next_row(ws, schema)

    is_new_row = ws.cell(row=row, column=_col(cols["item"])).value in (None, "")
    if is_new_row:
        ws.cell(row=row, column=_col(cols["item"])).value = last_item_value + 1
        total_col, minutos_col, fin_col, inicio_col = (
            cols["tiempo_total"],
            cols["tiempo_minutos"],
            cols["fin"],
            cols["inicio"],
        )
        ws[f"{total_col}{row}"] = f"={fin_col}{row}-{inicio_col}{row}"
        ws[f"{minutos_col}{row}"] = f"={total_col}{row}*24*60"
        ws[f"{cols['tiempo_horas']}{row}"] = f"={minutos_col}{row}/60"

    ws.cell(row=row, column=_col(cols["device"])).value = payload.dispositivo.strip()
    ws.cell(row=row, column=_col(cols["analista"])).value = payload.analista.strip()
    ws.cell(row=row, column=_col(cols["ticket"])).value = payload.ticket.strip()
    ws.cell(row=row, column=_col(cols["descripcion"])).value = (payload.descripcion or "").strip()
    ws.cell(row=row, column=_col(cols["inicio"])).value = payload.inicio
    ws.cell(row=row, column=_col(cols["causa"])).value = payload.causa.strip()
    ws.cell(row=row, column=_col(cols["areas"])).value = payload.areas.strip()
    ws.cell(row=row, column=_col(cols["solucion"])).value = payload.solucion.strip()
    ws.cell(row=row, column=_col(cols["fin"])).value = payload.fin

    if schema.has_tigo:
        ws.cell(row=row, column=_col(cols["acceso_tigo"])).value = (payload.acceso_tigo or "").strip()
        ws.cell(row=row, column=_col(cols["ticket_tigo_une"])).value = (
            payload.ticket_tigo_une or ""
        ).strip()

    return row


def compute_downtime_hours(ws, schema: PlatformSchema) -> tuple[float, int]:
    """Recalcula en Python (sin depender de formulas de Excel) el total de
    horas caidas registradas en la hoja, y cuantos eventos se contaron."""
    cols = schema.columns
    inicio_col = _col(cols["inicio"])
    fin_col = _col(cols["fin"])

    total_hours = 0.0
    count = 0
    for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + MAX_ROW_SCAN):
        inicio = ws.cell(row=row, column=inicio_col).value
        fin = ws.cell(row=row, column=fin_col).value
        if inicio is None and fin is None and ws.cell(row=row, column=1).value is None:
            break
        if isinstance(inicio, datetime) and isinstance(fin, datetime) and fin >= inicio:
            total_hours += (fin - inicio).total_seconds() / 3600
            count += 1
    return total_hours, count


def compute_disponibilidad(total_horas_mes: float, downtime_hours: float) -> float:
    if total_horas_mes <= 0:
        return 0.0
    return max(0.0, (total_horas_mes - downtime_hours) / total_horas_mes * 100)
