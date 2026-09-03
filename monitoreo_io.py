"""Creacion, carga y escritura del libro de monitoreo de disponibilidad
(equivalente en Excel al avail.csv de Analisis-Python.1), y calculo de la
lista de dispositivos caidos (por debajo de la meta de SLA), excluyendo los
marcados como "no controlados" (p. ej. fallas electricas) igual que en la
revision de la Planilla S@MI."""

import openpyxl
from openpyxl.styles import Font

from monitoreo_schema import COLUMNAS, ENCABEZADOS, HOJA_DISPONIBILIDAD, DispositivoPayload
from review import sugerir_no_controlado

FUENTE = "Arial"


def nuevo_libro_monitoreo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = HOJA_DISPONIBILIDAD
    for col_idx, encabezado in enumerate(ENCABEZADOS, start=1):
        celda = ws.cell(row=1, column=col_idx, value=encabezado)
        celda.font = Font(name=FUENTE, bold=True)
    ws.freeze_panes = "A2"
    return wb


def cargar_libro_monitoreo(path_or_buffer):
    wb = openpyxl.load_workbook(path_or_buffer, data_only=False)
    if HOJA_DISPONIBILIDAD not in wb.sheetnames:
        raise ValueError(
            f"El archivo no tiene una hoja '{HOJA_DISPONIBILIDAD}'. "
            "¿Es un archivo de monitoreo generado por esta herramienta?"
        )
    return wb


def agregar_dispositivo(wb, payload: DispositivoPayload) -> int:
    ws = wb[HOJA_DISPONIBILIDAD]
    row = ws.max_row + 1
    valores = {
        COLUMNAS["categoria"]: payload.categoria.strip(),
        COLUMNAS["host_name"]: payload.host_name.strip(),
        COLUMNAS["causa"]: payload.causa.strip(),
        COLUMNAS["no_controlado"]: "SI" if payload.no_controlado else "NO",
        COLUMNAS["percent_up"]: payload.percent_up,
        COLUMNAS["percent_down"]: payload.percent_down,
        COLUMNAS["percent_unreachable"]: payload.percent_unreachable,
        COLUMNAS["percent_undetermined"]: payload.percent_undetermined,
        COLUMNAS["total_time_down"]: payload.total_time_down,
    }
    for col_idx, encabezado in enumerate(ENCABEZADOS, start=1):
        clave = next(k for k, v in COLUMNAS.items() if v == encabezado)
        celda = ws.cell(row=row, column=col_idx, value=valores[COLUMNAS[clave]])
        celda.font = Font(name=FUENTE)
    return row


def leer_dispositivos(wb) -> list[dict]:
    ws = wb[HOJA_DISPONIBILIDAD]
    encabezados = [c.value for c in ws[1]]
    dispositivos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        fila = dict(zip(encabezados, row))
        dispositivos.append(fila)
    return dispositivos


def dispositivos_caidos(dispositivos: list[dict], umbral: float) -> dict:
    """Separa, igual que en la revision S@MI, el conteo 'total' (todas las
    caidas cuentan) del 'ajustado' (excluye las marcadas como no
    controladas, por ejemplo fallas electricas)."""
    caidos_total = [
        d for d in dispositivos
        if isinstance(d.get(COLUMNAS["percent_up"]), (int, float))
        and d[COLUMNAS["percent_up"]] < umbral
    ]
    caidos_ajustado = [
        d for d in caidos_total if str(d.get(COLUMNAS["no_controlado"], "")).upper() != "SI"
    ]
    return {
        "total": caidos_total,
        "ajustado": caidos_ajustado,
        "excluidos": [d for d in caidos_total if d not in caidos_ajustado],
    }


def sugerir_no_controlado_dispositivo(causa: str) -> bool:
    return sugerir_no_controlado(causa)
