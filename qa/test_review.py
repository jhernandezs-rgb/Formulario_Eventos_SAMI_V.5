import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parents[1]))

from review import (
    FilaRevisada,
    aplicar_correcciones,
    calcular_sla,
    nota_fin_de_semana,
    scan_platform,
    sugerir_no_controlado,
)
from schema import PLATFORMS
from workbook_io import load_workbook_for_write

SOURCE = Path(__file__).parents[1] / "sami_source_reparado.xlsx"


def _fresh_workbook():
    return load_workbook_for_write(SOURCE)


def _fila(**overrides) -> FilaRevisada:
    base = dict(
        fila=3,
        item=1,
        dispositivo="Sw_Test",
        analista="Julian / 06:00 - 14:00",
        ticket="IM-0001",
        causa="Bloqueo de Sw",
        areas="TELECOMUNICACIONES",
        solucion="Reinicio",
        inicio=datetime(2026, 7, 1, 8, 0),
        fin=datetime(2026, 7, 1, 8, 30),
        duracion_horas=0.5,
    )
    base.update(overrides)
    return FilaRevisada(**base)


class TestSugerirNoControlado:
    def test_causa_electrica_se_sugiere_no_controlada(self):
        assert sugerir_no_controlado("Falla electrica en el rack") is True

    def test_causa_proveedor_tigo_se_sugiere_no_controlada(self):
        assert sugerir_no_controlado("Retiro del servicio Tigo") is True

    def test_causa_interna_no_se_sugiere(self):
        assert sugerir_no_controlado("Bloqueo de Sw") is False

    def test_causa_vacia_no_se_sugiere(self):
        assert sugerir_no_controlado("") is False


class TestNotaFinDeSemana:
    def test_sabado_genera_nota(self):
        # 2026-07-04 es sabado.
        nota = nota_fin_de_semana(datetime(2026, 7, 4, 10, 0))
        assert nota is not None
        assert "sabado" in nota
        assert "no se descuenta tiempo" in nota

    def test_domingo_genera_nota(self):
        # 2026-07-05 es domingo.
        nota = nota_fin_de_semana(datetime(2026, 7, 5, 10, 0))
        assert nota is not None
        assert "domingo" in nota

    def test_dia_habil_no_genera_nota(self):
        # 2026-07-01 es miercoles.
        assert nota_fin_de_semana(datetime(2026, 7, 1, 10, 0)) is None

    def test_none_no_genera_nota(self):
        assert nota_fin_de_semana(None) is None

    def test_incidente_fin_de_semana_no_reduce_duracion(self):
        """La nota es solo informativa: la duracion se calcula igual, sin
        descontar tiempo por caer en fin de semana."""
        fila = _fila(
            fila=3,
            inicio=datetime(2026, 7, 4, 8, 0),  # sabado
            fin=datetime(2026, 7, 6, 8, 0),  # lunes, 48 horas despues
            duracion_horas=48.0,
        )
        resultado = calcular_sla([fila], total_horas_mes=1000, no_controlados=set())
        assert resultado["horas_totales_caida"] == 48.0


class TestScanPlatform:
    def test_telecomunicaciones_detecta_filas_reales(self):
        wb = _fresh_workbook()
        schema = PLATFORMS["Telecomunicaciones"]
        filas = scan_platform(wb[schema.sheet], schema)
        assert len(filas) > 0
        # Las filas placeholder totalmente vacias no deben contarse.
        assert all(f.dispositivo or f.analista or f.inicio for f in filas)

    def test_fin_anterior_a_inicio_se_marca_como_problema(self):
        wb = _fresh_workbook()
        schema = PLATFORMS["SAP"]
        ws = wb[schema.sheet]
        cols = schema.columns
        from openpyxl.utils import column_index_from_string

        def col(letter):
            return column_index_from_string(letter)

        ws.cell(row=3, column=col(cols["device"])).value = "Servidor SAP"
        ws.cell(row=3, column=col(cols["analista"])).value = "Julian"
        ws.cell(row=3, column=col(cols["ticket"])).value = "IM-1"
        ws.cell(row=3, column=col(cols["causa"])).value = "Bloqueo"
        ws.cell(row=3, column=col(cols["areas"])).value = "SERVIDORES"
        ws.cell(row=3, column=col(cols["solucion"])).value = "Reinicio"
        ws.cell(row=3, column=col(cols["inicio"])).value = datetime(2026, 7, 1, 14, 0)
        ws.cell(row=3, column=col(cols["fin"])).value = datetime(2026, 7, 1, 2, 0)

        filas = scan_platform(ws, schema)
        fila = next(f for f in filas if f.fila == 3)
        assert any("Fin anterior a Inicio" in p for p in fila.problemas)

    def test_campos_vacios_se_marcan_como_problema(self):
        wb = _fresh_workbook()
        schema = PLATFORMS["SAP"]
        ws = wb[schema.sheet]
        cols = schema.columns
        from openpyxl.utils import column_index_from_string

        ws.cell(row=3, column=column_index_from_string(cols["device"])).value = "Servidor SAP"
        # Analista, ticket, causa, areas, solucion, inicio y fin quedan vacios.

        filas = scan_platform(ws, schema)
        fila = next(f for f in filas if f.fila == 3)
        assert "Analista vacio" in fila.problemas
        assert "Numero de Ticket vacio" in fila.problemas
        assert "Causa vacia" in fila.problemas


class TestCalcularSla:
    def test_sin_exclusiones_total_y_ajustada_son_iguales(self):
        filas = [_fila(fila=3, duracion_horas=10.0)]
        resultado = calcular_sla(filas, total_horas_mes=1000, no_controlados=set())
        assert resultado["disponibilidad_total"] == resultado["disponibilidad_ajustada"]
        assert resultado["disponibilidad_total"] == 99.0

    def test_excluir_fila_mejora_la_disponibilidad_ajustada(self):
        filas = [
            _fila(fila=3, duracion_horas=10.0, causa="Bloqueo de Sw"),
            _fila(fila=4, duracion_horas=20.0, causa="Falla electrica"),
        ]
        resultado = calcular_sla(filas, total_horas_mes=1000, no_controlados={4})
        assert resultado["horas_no_controladas"] == 20.0
        assert resultado["horas_controladas"] == 10.0
        assert resultado["disponibilidad_ajustada"] > resultado["disponibilidad_total"]


class TestAplicarCorrecciones:
    def test_correcciones_se_escriben_en_la_hoja(self):
        wb = _fresh_workbook()
        schema = PLATFORMS["SAP"]
        ws = wb[schema.sheet]
        filas = scan_platform(ws, schema)
        if not filas:
            filas = [_fila(fila=3)]
        filas[0].dispositivo = "Servidor Corregido"
        filas[0].analista = "Nuevo Analista"

        aplicar_correcciones(ws, schema, filas)

        from openpyxl.utils import column_index_from_string

        cols = schema.columns
        assert (
            ws.cell(row=filas[0].fila, column=column_index_from_string(cols["device"])).value
            == "Servidor Corregido"
        )
        assert (
            ws.cell(row=filas[0].fila, column=column_index_from_string(cols["analista"])).value
            == "Nuevo Analista"
        )


if __name__ == "__main__":
    import pytest

    pytest.main([__file__, "-v"])
