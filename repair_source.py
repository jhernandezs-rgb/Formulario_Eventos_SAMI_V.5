"""Repara errores conocidos del libro S@MI para poder abrirlo en modo
escritura y capturar eventos nuevos sin arrastrar errores de formula.

Corrige:
1. La relacion de imagen rota (Target="NULL") en xl/drawings/_rels/drawing1.xml.rels
   que hace que openpyxl falle con KeyError al abrir en modo escritura.
2. Hoja "Data Center": el encabezado de la columna F (fila 2) quedo sobrescrito
   con ":40" en vez de "Fecha/Hora Inicio Evento ", y la formula de Tiempo
   Total de la fila 3 quedo como "=#REF!-J3" en vez de "=J3-F3" (unica fila
   distinta del patron que usan el resto de filas de esa misma columna).

Uso:
    py repair_source.py "<ruta origen.xlsx>" "<ruta destino.xlsx>"
"""
import sys
import warnings
import zipfile

import openpyxl

REL_PATH = "xl/drawings/_rels/drawing1.xml.rels"
BROKEN_TARGET = 'Target="NULL"'
FIXED_TARGET = 'Target="../media/image1.jpeg"'


def _repair_image_relationship(src: str, dst: str) -> None:
    with zipfile.ZipFile(src, "r") as zin:
        names = zin.namelist()
        rels_xml = zin.read(REL_PATH).decode("utf-8") if REL_PATH in names else ""
        with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == REL_PATH and BROKEN_TARGET in rels_xml:
                    data = rels_xml.replace(BROKEN_TARGET, FIXED_TARGET).encode("utf-8")
                zout.writestr(item, data)


def _repair_data_center(dst: str) -> None:
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(dst, data_only=False)
    ws = wb["Data Center"]
    if ws["F2"].value == ":40":
        ws["F2"] = "Fecha/Hora Inicio Evento "
    if ws["K3"].value == "=#REF!-J3":
        ws["K3"] = "=J3-F3"
    wb.save(dst)


def repair(src: str, dst: str) -> None:
    _repair_image_relationship(src, dst)
    print(f"Reparado: {REL_PATH} -> rId3 ahora apunta a image1.jpeg")
    _repair_data_center(dst)
    print("Reparado: Data Center!F2 (encabezado) y Data Center!K3 (formula Tiempo Total)")
    print(f"Copia reparada guardada en: {dst}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        raise SystemExit(1)
    repair(sys.argv[1], sys.argv[2])
