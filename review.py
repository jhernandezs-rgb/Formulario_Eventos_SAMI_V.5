"""Logica de revision del archivo mensual ya descargado: recorre cada hoja
de plataforma, detecta problemas de captura (campos vacios, Fin anterior a
Inicio, duraciones absurdas que delatan una hora mal digitada) y sugiere que
incidencias son "no controladas" (causas externas, tipicamente fallas
electricas o de proveedor) para poder calcular una disponibilidad ajustada
que las excluye, ademas de la disponibilidad total sin ajustar.

La clasificacion "controlado / no controlado" es siempre una sugerencia:
el usuario la revisa y la puede cambiar fila por fila antes de calcular el
SLA final."""

from dataclasses import dataclass, field
from datetime import datetime

from openpyxl.utils import column_index_from_string

from schema import FIRST_DATA_ROW, PlatformSchema
from workbook_io import clean_analista_name

MAX_ROW_SCAN = 5000
DURACION_MAXIMA_HORAS = 24 * 31

# Palabras clave (en minuscula) que sugieren una causa externa / no
# controlada por el equipo de TI. Es solo un punto de partida editable por
# el revisor, no una regla automatica definitiva.
PALABRAS_NO_CONTROLADO = [
    "electric",
    "eléctric",
    "energia",
    "energía",
    "fluido electrico",
    "fluido eléctrico",
    "corte de energia",
    "corte de energía",
    "tigo",
    "une",
    "proveedor",
    "tercero",
    "mantenimiento programado",
]


def sugerir_no_controlado(causa: str) -> bool:
    if not causa:
        return False
    texto = causa.lower()
    return any(palabra in texto for palabra in PALABRAS_NO_CONTROLADO)


@dataclass
class FilaRevisada:
    fila: int
    item: object
    dispositivo: str
    analista: str
    ticket: str
    causa: str
    areas: str
    solucion: str
    inicio: datetime | None
    fin: datetime | None
    duracion_horas: float
    problemas: list[str] = field(default_factory=list)
    notas: list[str] = field(default_factory=list)
    no_controlado_sugerido: bool = False


def nota_fin_de_semana(inicio: datetime | None) -> str | None:
    """Las fallas que inician en fin de semana SI cuentan completas contra
    el SLA (no se les resta tiempo); solo se deja constancia de que se
    evaluaron hasta el primer dia habil siguiente, para que quede
    documentado por que la atencion no fue inmediata."""
    if inicio is None:
        return None
    if inicio.weekday() in (5, 6):  # 5 = sabado, 6 = domingo
        dia = "sabado" if inicio.weekday() == 5 else "domingo"
        return (
            f"Incidente iniciado en fin de semana ({dia}): se evalua hasta el "
            "primer dia habil siguiente. Las horas de caida cuentan completas, "
            "no se descuenta tiempo."
        )
    return None


def _col(letter: str) -> int:
    return column_index_from_string(letter)


def _valida_fila(schema: PlatformSchema, ws, row: int) -> FilaRevisada | None:
    cols = schema.columns
    item = ws.cell(row=row, column=_col(cols["item"])).value
    if item in (None, ""):
        return None

    dispositivo = ws.cell(row=row, column=_col(cols["device"])).value
    analista = ws.cell(row=row, column=_col(cols["analista"])).value
    ticket = ws.cell(row=row, column=_col(cols["ticket"])).value
    causa = ws.cell(row=row, column=_col(cols["causa"])).value
    areas = ws.cell(row=row, column=_col(cols["areas"])).value
    solucion = ws.cell(row=row, column=_col(cols["solucion"])).value
    inicio = ws.cell(row=row, column=_col(cols["inicio"])).value
    fin = ws.cell(row=row, column=_col(cols["fin"])).value

    # Fila "placeholder" sin usar: el Item ya viene precargado en toda la
    # hoja, pero si no hay ningun otro dato, no es un evento real.
    if dispositivo in (None, "") and analista in (None, "") and inicio is None and fin is None:
        return None

    problemas: list[str] = []
    if not dispositivo:
        problemas.append(f"{schema.device_label} vacio")
    if not analista:
        problemas.append("Analista vacio")
    if not ticket:
        problemas.append("Numero de Ticket vacio")
    if not causa:
        problemas.append("Causa vacia")
    if not areas:
        problemas.append("Areas Involucradas vacio")
    if not solucion:
        problemas.append("Solucion vacia")
    if inicio is None:
        problemas.append("Fecha/Hora Inicio vacia")
    if fin is None:
        problemas.append("Fecha/Hora Fin vacia")

    duracion_horas = 0.0
    if isinstance(inicio, datetime) and isinstance(fin, datetime):
        if fin < inicio:
            problemas.append(
                "Fin anterior a Inicio: revisa si la hora se digito en el "
                "formato equivocado (24h vs 12h)"
            )
        else:
            duracion_horas = (fin - inicio).total_seconds() / 3600
            if duracion_horas > DURACION_MAXIMA_HORAS:
                problemas.append(
                    f"Duracion de {duracion_horas:.1f} h: revisa si la hora "
                    "se digito en el formato equivocado (24h vs 12h)"
                )

    inicio_dt = inicio if isinstance(inicio, datetime) else None
    notas = []
    nota_finde = nota_fin_de_semana(inicio_dt)
    if nota_finde:
        notas.append(nota_finde)

    return FilaRevisada(
        fila=row,
        item=item,
        dispositivo=dispositivo or "",
        analista=clean_analista_name(analista or ""),
        ticket=ticket or "",
        causa=causa or "",
        areas=areas or "",
        solucion=solucion or "",
        inicio=inicio_dt,
        fin=fin if isinstance(fin, datetime) else None,
        duracion_horas=duracion_horas,
        problemas=problemas,
        notas=notas,
        no_controlado_sugerido=sugerir_no_controlado(causa or ""),
    )


def scan_platform(ws, schema: PlatformSchema) -> list[FilaRevisada]:
    """Recorre todas las filas con datos de la hoja y las devuelve
    analizadas. No modifica el libro."""
    filas: list[FilaRevisada] = []
    for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + MAX_ROW_SCAN):
        revisada = _valida_fila(schema, ws, row)
        if revisada is None and ws.cell(row=row, column=1).value in (None, ""):
            break
        if revisada is not None:
            filas.append(revisada)
    return filas


def calcular_sla(
    filas: list[FilaRevisada], total_horas_mes: float, no_controlados: set[int]
) -> dict:
    """Calcula disponibilidad total (todas las horas de caida cuentan) y
    disponibilidad ajustada (excluyendo las filas marcadas como no
    controladas). `no_controlados` es un set de numeros de fila."""
    horas_totales = sum(f.duracion_horas for f in filas)
    horas_no_controladas = sum(
        f.duracion_horas for f in filas if f.fila in no_controlados
    )
    horas_controladas = horas_totales - horas_no_controladas

    def disponibilidad(horas_caida: float) -> float:
        if total_horas_mes <= 0:
            return 0.0
        return max(0.0, (total_horas_mes - horas_caida) / total_horas_mes * 100)

    return {
        "horas_totales_caida": horas_totales,
        "horas_no_controladas": horas_no_controladas,
        "horas_controladas": horas_controladas,
        "disponibilidad_total": disponibilidad(horas_totales),
        "disponibilidad_ajustada": disponibilidad(horas_controladas),
    }


def aplicar_correcciones(ws, schema: PlatformSchema, filas: list[FilaRevisada]) -> None:
    """Escribe de vuelta en la hoja los valores corregidos por el revisor:
    Dispositivo/Elemento, Analista, Ticket, Causa, Areas Involucradas,
    Solucion, y Fecha/Hora Inicio y Fin (estas ultimas se editan con un
    selector de fecha/hora en la tabla, nunca como texto libre, para no
    reintroducir el problema de formato 24h/12h que origino esta revision)."""
    cols = schema.columns
    for f in filas:
        ws.cell(row=f.fila, column=_col(cols["device"])).value = f.dispositivo
        ws.cell(row=f.fila, column=_col(cols["analista"])).value = f.analista
        ws.cell(row=f.fila, column=_col(cols["ticket"])).value = f.ticket
        ws.cell(row=f.fila, column=_col(cols["causa"])).value = f.causa
        ws.cell(row=f.fila, column=_col(cols["areas"])).value = f.areas
        ws.cell(row=f.fila, column=_col(cols["solucion"])).value = f.solucion
        if f.inicio is not None:
            ws.cell(row=f.fila, column=_col(cols["inicio"])).value = f.inicio
        if f.fin is not None:
            ws.cell(row=f.fila, column=_col(cols["fin"])).value = f.fin
